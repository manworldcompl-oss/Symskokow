
# excel_saver.py
# Jedno źródło prawdy dla zapisu XLSX w projekcie Ski Jump Simulator.
# Używa openpyxl; dba o sanitację danych, formaty i szerokości kolumn.

from __future__ import annotations
from pathlib import Path
from typing import Dict, Iterable
import re
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# --- Sanitizer podobny do _sanitize_df_STRONG ---
def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    import numpy as np
    import re, math
    if df is None:
        return pd.DataFrame()
    try:
        out = pd.DataFrame(df).copy()
    except Exception:
        return pd.DataFrame()

    # wyczyść nagłówki z kontrolnych i przytnij
    def _clean_header(x: str) -> str:
        s = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", str(x))
        return s[:255]
    out.columns = [_clean_header(c) for c in out.columns]

    # inf -> NaN
    num_cols = out.select_dtypes(include=["number"]).columns
    if len(num_cols):
        out[num_cols] = out[num_cols].replace([np.inf, -np.inf], pd.NA)

    # sanitizer komórki (tylko dla kolumn tekstowych)
    def _cell(x):
        try:
            if x is None or (isinstance(x, float) and math.isnan(x)):
                return ""
        except Exception:
            pass
        if isinstance(x, (bytes, bytearray)):
            try:
                x = x.decode("utf-8", errors="replace")
            except Exception:
                x = str(x)
        if isinstance(x, (list, tuple, set, dict)):
            x = str(x)
        s = str(x)
        s = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", s)
        return s[:32760]  # limit Excela 32767

    # KLUCZ: iteruj po indeksach, żeby obsłużyć duplikaty nazw kolumn
    for i in range(out.shape[1]):
        s = out.iloc[:, i]
        dt = s.dtype
        if dt == "object" or str(dt).startswith(("string", "category")):
            out.iloc[:, i] = s.map(_cell)

    return out

# heurystyki formatowania liczb wg nazwy kolumny
_ONE_DEC_COLS = {"Punkty","Punkty rundy","Punkty za odległość","Noty stylowe","Kompensacja wiatru","Kompensacja belki","PTS"}
_ZERO_DEC_COLS = {"Seed","Lp.","LP.","Miejsce","Para"}

def _guess_format(col: str) -> str | None:
    u = str(col).strip().upper()
    if col in _ZERO_DEC_COLS or u in {"LP", "LP."}:
        return "0"
    if col in _ONE_DEC_COLS or "ODL" in u or "ODLEG" in u:
        return "0.0"
    return None  # tekst domyślnie

def _autosize(ws):
    # proste autosize: maks długość tekstu w kolumnie + margines
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            try:
                v = cell.value
                length = len(str(v)) if v is not None else 0
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(80, max(8, max_len + 2))

def _apply_header_style(ws):
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    bold = Font(bold=True)
    center = Alignment(vertical="center")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(top=thin, bottom=thin)
    first = True
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = center
        cell.border = border
    ws.freeze_panes = "A2"
    # autofilter
    ws.auto_filter.ref = ws.dimensions

def _apply_number_formats(ws):
    for col_idx, cell in enumerate(ws[1], start=1):
        fmt = _guess_format(cell.value)
        if fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col_idx).number_format = fmt

def save_competition_results(filename: str | Path, sheets: dict[str, pd.DataFrame]) -> str:
    """
    Zapisuje wieloarkuszowy XLSX z ujednoliconym stylem.
    Zwraca ścieżkę do pliku.
    """
    p = Path(filename)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # usuwamy domyślny arkusz
    wb.remove(wb.active)

    for name, df in sheets.items():
        ws = wb.create_sheet(title=str(name)[:31] or "Arkusz")
        sdf = _sanitize_df(df)
        for r in dataframe_to_rows(sdf, index=False, header=True):
            ws.append(r)
        _apply_header_style(ws)
        _apply_number_formats(ws)
        _autosize(ws)

    wb.save(p)
    return str(p.resolve())
